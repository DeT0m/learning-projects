#!usr/bin/env python3
# multiplicationTable.py - Creates excel spreadsheet with multiplication
# table for a given number

import sys
import openpyxl
from openpyxl.styles import Font

number = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
bold = Font(bold=True, name='Calibri')

for i in range(1, number + 1):
    row = sheet.cell(row=1, column=i + 1)
    row.value = i
    row.font = bold
    col = sheet.cell(row=i + 1, column=1)
    col.value = i
    col.font = bold
    for i in range(1, number + 1):
        sheet.cell(row=i+1, column=col.value + 1).value = row.value * i

wb.save('multiTable.xlsx')
