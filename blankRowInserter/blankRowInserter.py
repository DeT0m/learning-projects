#!usr/bin/env python3
# blankRowInserter.py - Inserts given number of empty rows in Excel spreadsheet
# after the specified row number.

import openpyxl
from openpyxl.utils import get_column_letter
import sys

n_number = int(sys.argv[1])  # Number of rows to insert
m_number = int(sys.argv[2])  # Insert empty rows after this row number
excel_file = sys.argv[3]  # Excel filename

# Loading workbook
wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

# Defining range of cells to move
new_range = (
    'A'
    + str(m_number + 1)
    + ':'
    + str(get_column_letter(sheet.max_column))
    + str(sheet.max_row)
    )

# Moving range cell
sheet.move_range(new_range, rows=n_number, translate=True)

# Saving file
wb.save(excel_file)
