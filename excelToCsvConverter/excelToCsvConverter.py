#!usr/bin/env python3
# excelToCsvConverter.py - Reads contents of Excel file and saves it as
# csv files.

import os
import openpyxl
import csv
from pathlib import Path

os.makedirs('exported', exist_ok=True)

for excelFile in os.listdir(Path.cwd()):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(excelFile)

    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb[sheetName]

        # Create the CSV filename from the Excel filename and sheet title.
        csvName = (Path(os.path.join(Path.cwd(), excelFile)).stem
                   + '_'
                   + sheetName
                   + '.csv')
        csvFile = open((Path('exported') / Path(csvName)), 'w')
        # Create the csv.writer object for this CSV file.
        csvWriter = csv.writer(csvFile)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)
            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)
        csvFile.close()
