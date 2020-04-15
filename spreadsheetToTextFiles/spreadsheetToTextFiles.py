#!usr/bin/env python3
# spreadsheetToTextFiles.py - Get all lines in column from spreadsheet and
# save it as text file. One file per column.

import openpyxl
import sys

file = sys.argv[1]  # Get filenames from CL arguments

# Create spreadsheet
wb = openpyxl.load_workbook(file)
sheet = wb.active

for i in range(1, sheet.max_column + 1):  # Iterate over columns in file
    col_num = i  # Get column number (file number), starting at 1
    file = open('file' + str(i) + '.txt', 'w')  # Create new file
    for j in range(1, sheet.max_row + 1):  # Iterate over rows in column
        row_num = j  # Get row number (line number), starting at 1
        value = sheet.cell(column=col_num, row=j).value  # Get value from cell
        # Save every line from from cell to file
        if value is not None:
            file.write(value)
    file.close()  # Close file
