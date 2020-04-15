#!usr/bin/env python3
# spreadsheetCellInverter.py - Inverts positions of cells in Excel spreadsheet
# and saves it to new worksheet

import openpyxl
import sys

filename = sys.argv[1]  # Get filename
wb = openpyxl.load_workbook(filename)
sheet = wb.active
# Create new spreadsheet
wb.create_sheet(title='Transposed')
new_sheet = wb['Transposed']

position = []  # List for columns

for i in range(1, sheet.max_column + 1):  # Iterate over columns
    col = i
    position.append([])  # Adding list for row numbers
    for i in range(1, sheet.max_row):  # Iterate over rows
        position[col - 1].append(i)  # Adding individual row numbers to list

for col in range(len(position)):  # Iterate over columns
    new_row = col + 1  # Get new row number
    for row in range(len(position[col]) + 1):  # Iterate over rows
        new_col = row + 1  # Get new column number
        # Assign values to new cell positions
        new_sheet.cell(row=new_row, column=new_col).value = sheet.cell(
            row=new_col, column=new_row).value

wb.save(filename)
